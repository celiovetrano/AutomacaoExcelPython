import re
from openpyxl import Workbook, load_workbook
from models.inicializacao import Inicializacao
from datetime import datetime
import json

class CadastroFuncionario:
    # Função para obter RF no formato correto
    def obter_rf(self):
        while True:
            rf = input("Digite o RF (formato 000.000.0.00/0): ")
            if re.match(r"^\d{3}\.\d{3}\.\d\.\d{2}/\d$", rf):
                return rf
            else:
                print("RF inválido. Por favor, insira no formato 000.000.0.00/0.")

    # Função para carregar dados dos funcionários
    def carregar_funcionarios(self):
        try:
            wb = load_workbook("servidor.xlsx")
            ws = wb.active
            funcionarios = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and len(row) == 8:  # Garantir que a linha não esteja vazia e tenha 8 colunas
                    horarios_jeif = json.loads(row[4].replace("'", '"'))  # Substituir aspas simples por duplas
                    horarios_regencia = json.loads(row[5].replace("'", '"'))  # Substituir aspas simples por duplas
                    funcionarios[row[0]] = [row[1], row[2], row[3], horarios_jeif, horarios_regencia, row[6], row[7]]
            return funcionarios
        except FileNotFoundError:
            return {}

    # Função para salvar dados dos funcionários
    def salvar_funcionarios(self, funcionarios):
        wb = Workbook()
        ws = wb.active
        ws.append(["RF", "Nome", "QPE", "Início do Exercício", "Horários JEIF", "Horários de Regência", "Série de Regência", "Cargo"])
        for rf, dados in funcionarios.items():
            if len(dados) == 7:  # Garantir que todos os dados estão presentes
                # Serializar os horários como strings para salvar no Excel
                horarios_jeif = json.dumps(dados[3])
                horarios_regencia = json.dumps(dados[4])
                ws.append([rf, dados[0], dados[1], dados[2], horarios_jeif, horarios_regencia, dados[5], dados[6]])
        wb.save("servidor.xlsx")

    # Função para obter horários diários
    def obter_horarios_diarios(self, tipo):
        horarios = {}
        for dia in ["segunda-feira", "terça-feira", "quarta-feira", "quinta-feira", "sexta-feira"]:
            while True:
                inicio = input(f"Digite o horário de início de {tipo} na {dia} (HH:MM): ").strip()
                fim = input(f"Digite o horário de fim de {tipo} na {dia} (HH:MM): ").strip()
                if self.validar_horario(inicio) and self.validar_horario(fim):
                    horarios[dia] = f"{inicio} às {fim}"
                    break
                else:
                    print("Formato de horário inválido. Por favor, insira no formato HH:MM.")
        return horarios

    # Função para validar o formato de horário
    def validar_horario(self, horario):
        try:
            datetime.strptime(horario, '%H:%M')
            return True
        except ValueError:
            return False

    # Função para adicionar um funcionário
    def adicionar_funcionario(self, funcionarios):
        while True:
            Inicializacao.limpar_tela()
            rf = self.obter_rf()
            nome = input("Digite o nome do funcionário: ").strip()
            qpe = input("Digite o QPE: ").strip()
            while True:
                inicio_exercicio = input("Digite a data de início do exercício (formato DD/MM/AAAA): ").strip()
                if self.validar_data(inicio_exercicio):
                    break
                else:
                    print("Formato de data inválido. Por favor, insira no formato DD/MM/AAAA.")
            horarios_jeif = self.obter_horarios_diarios("JEIF")
            horarios_regencia = self.obter_horarios_diarios("regência")
            serie_regencia = input("Digite a série de regência (formato número e letra, ex.: 3A): ").strip()
            cargo = input("Digite o cargo do funcionário: ").strip()
            
            funcionarios[rf] = [nome, qpe, inicio_exercicio, horarios_jeif, horarios_regencia, serie_regencia, cargo]
            self.salvar_funcionarios(funcionarios)
            print(f"Funcionário {nome} cadastrado com sucesso!")
        
            if input("Pressione 'Espaço' para retornar ao menu: ") == ' ':
                break
        
        Inicializacao.limpar_tela()
        Inicializacao.exibir_nome_aplicaçao()
        Inicializacao.menu()

    # Função para validar o formato de data
    def validar_data(self, data):
        try:
            datetime.strptime(data, '%d/%m/%Y')
            return True
        except ValueError:
            return False

    # Função para consultar funcionários
    def consultar_funcionarios(self, funcionarios):
        while True:
            Inicializacao.limpar_tela()
            if funcionarios:
                for rf, dados in funcionarios.items():
                    if len(dados) == 7:
                        nome, qpe, inicio_exercicio, horarios_jeif, horarios_regencia, serie_regencia, cargo = dados
                        print(f"RF: {rf}\nNome: {nome}\nQPE: {qpe}\nInício do Exercício: {inicio_exercicio}\nHorários JEIF:\n{horarios_jeif}\nHorários de Regência:\n{horarios_regencia}\nSérie de Regência: {serie_regencia}\nCargo: {cargo}\n")
                    else:
                        print(f"Dados incompletos para o funcionário com RF: {rf}")
            else:
                print("Nenhum funcionário cadastrado.")

            if input("Pressione 'Espaço' para retornar ao menu: ") == ' ':
                break

        Inicializacao.limpar_tela()
        Inicializacao.exibir_nome_aplicaçao()
        Inicializacao.menu()

    # Função para alterar dados de um funcionário
    def alterar_funcionario(self, funcionarios):
        while True:
            Inicializacao.limpar_tela()
            rf = self.obter_rf()
            if rf in funcionarios:
                nome = input(f"Digite o novo nome para o RF {rf}: ").strip()
                qpe = input(f"Digite o novo QPE para o RF {rf}: ").strip()
                while True:
                    inicio_exercicio = input(f"Digite a nova data de início do exercício para o RF {rf} (formato DD/MM/AAAA): ").strip()
                    if self.validar_data(inicio_exercicio):
                        break
                    else:
                        print("Formato de data inválido. Por favor, insira no formato DD/MM/AAAA.")
                horarios_jeif = self.obter_horarios_diarios("JEIF")
                horarios_regencia = self.obter_horarios_diarios("regência")
                serie_regencia = input(f"Digite a nova série de regência do RF {rf} (formato número e letra, ex.: 3A): ").strip()
                cargo = input(f"Digite o novo cargo do RF {rf}: ").strip()
                
                funcionarios[rf] = [nome, qpe, inicio_exercicio, horarios_jeif, horarios_regencia, serie_regencia, cargo]
                self.salvar_funcionarios(funcionarios)
                print(f"Dados do funcionário {nome} atualizados com sucesso!")
            else:
                print("Funcionário não encontrado.")

            if input("Pressione 'Espaço' para retornar ao menu: ") == ' ':
                break

        Inicializacao.limpar_tela()
        Inicializacao.exibir_nome_aplicaçao()
        Inicializacao.menu()

    # Função para excluir um funcionário
    def excluir_funcionario(self, funcionarios):
        while True:
            Inicializacao.limpar_tela()
            rf = self.obter_rf()
            if rf in funcionarios:
                del funcionarios[rf]
                self.salvar_funcionarios(funcionarios)
                print(f"Funcionário {rf} excluído com sucesso!")
            else:
                print("Funcionário não encontrado.")

            if input("Pressione 'Espaço' para retornar ao menu: ") == ' ':
                break

        Inicializacao.limpar_tela()
        Inicializacao.exibir_nome_aplicaçao()
        Inicializacao.menu()
