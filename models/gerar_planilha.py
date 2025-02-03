import os
import re
import calendar
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import locale
import datetime

class GerarPlanilha:
    # Função para obter o mês por extenso
    def obter_nome_mes(self, mes):
        # Configurações do usuário
        locale.setlocale(locale.LC_ALL, "")
        nome_mes = calendar.month_name[mes]
        return nome_mes

    # Função para recuperar primeiro e último dia do mês / ano
    def primeiro_e_ultimo_dia_do_mes(self, ano, mes):
        # Primeiro dia do mês
        primeiro_dia = datetime.date(ano, mes, 1)
        
        # Último dia do mês
        if mes == 12:
            ultimo_dia = datetime.date(ano + 1, 1, 1) - datetime.timedelta(days=1)
        else:
            ultimo_dia = datetime.date(ano, mes + 1, 1) - datetime.timedelta(days=1)
        
        # Formatar as datas para o padrão brasileiro
        primeiro_dia_formatado = primeiro_dia.strftime('%d/%m/%Y')
        ultimo_dia_formatado = ultimo_dia.strftime('%d/%m/%Y')
        
        return primeiro_dia_formatado, ultimo_dia_formatado

    # Função para validar datas do mês para definir se é final de semana ou está em período de recesso.
    def ajustar_data(self, mes, ano, ws, recesso_inicio=None, recesso_fim=None):
        # Define os dias da semana
        dias_semana_completo = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]
        dias_semana_inicial = ["S", "T", "Q", "Q", "S", "S", "D"]
        
        # Obtem o número de dias no mês e o primeiro dia da semana
        num_dias = calendar.monthrange(ano, mes)[1]
        
        # Define a linha inicial e a coluna inicial
        linha_inicial = 17
        coluna_inicial = 2
        
        for dia in range(1, 32):  # Verificamos até o dia 31
            if dia <= num_dias:
                dia_semana = calendar.weekday(ano, mes, dia)
                ws.cell(row=linha_inicial + dia - 1, column=coluna_inicial, value=dia)  # Preenche o dia
                ws.cell(row=linha_inicial + dia - 1, column=coluna_inicial + 1, value=dias_semana_inicial[dia_semana])  # Preenche a inicial do dia da semana

                # Verifica se o dia está dentro do período de recesso
                if recesso_inicio and recesso_fim and recesso_inicio <= dia <= recesso_fim:
                    # Preenche a coluna "P" com "Recesso" para os dias do recesso
                    if dia_semana in [5, 6]:  # Mantém a configuração atual para finais de semana
                        ws.cell(row=linha_inicial + dia - 1, column=16, value=dias_semana_completo[dia_semana])  # Coluna V é a coluna 16
                    else:
                        ws.cell(row=linha_inicial + dia - 1, column=16, value="Recesso")
                else:
                    # Preenche a coluna "P" com o nome do dia da semana por extenso apenas para finais de semana
                    if dia_semana in [5, 6]:
                        ws.cell(row=linha_inicial + dia - 1, column=16, value=dias_semana_completo[dia_semana])  # Coluna P é a coluna 16
                    else:
                        ws.cell(row=linha_inicial + dia - 1, column=16, value=None)  # Limpa a coluna P se não for sábado ou domingo

    def editar_aba(self, dre, escola, arquivo_origem, nome_nova_aba, mes, ano, nome_servidor, rf, qpe, inicio_exercicio, horarios_jeif, horarios_regencia, sala_regencia, cargo, diretorio, recesso_inicio=None, recesso_fim=None):
        # Abre o arquivo de origem com opção data_only
        wb_origem = load_workbook(arquivo_origem, keep_vba=True, data_only=True)

        # Seleciona a aba ativa
        aba_ativa = wb_origem.active

        # Copia a imagem existente na planilha de origem
        if 'A1' in aba_ativa:
            img = aba_ativa['A1']
            aba_ativa.add_image(img, 'A1')

        # Renomeia a aba ativa
        aba_ativa.title = nome_nova_aba.capitalize()

        # Faz as edições necessárias na aba ativa
        aba_ativa['A3'] = f"DIRETORIA REGIONAL DE EDUCAÇÃO DE {dre}".upper()
        aba_ativa['A4'] = escola.upper()
        aba_ativa['K8'] = f"  SEDE: {escola}".upper()
        aba_ativa['A5'] = nome_servidor  # Nome do servidor
        aba_ativa['A6'] = cargo  # Cargo do servidor
        aba_ativa['D7'] = rf  # RF
        aba_ativa['E7'] = f"QPE: {qpe}"  # QPE
        aba_ativa['I7'] = f"INÍCIO DO EXERCÍCIO: {inicio_exercicio}"  # Início do exercício
        aba_ativa['A11'] = sala_regencia  # Sala de regência

        # Horários JEIF e regência por dia da semana
        colunas_jeif = ['C10', 'E10', 'I10', 'N10', 'T10']
        colunas_regencia = ['C11', 'E11', 'I11', 'N11', 'T11']
        dias_semana = ["segunda-feira", "terça-feira", "quarta-feira", "quinta-feira", "sexta-feira"]

        for i, dia in enumerate(dias_semana):
            aba_ativa[colunas_jeif[i]] = horarios_jeif[dia]
            aba_ativa[colunas_regencia[i]] = horarios_regencia[dia]

        # Obter primeiro e último dia do mês
        primeiro_dia, ultimo_dia = self.primeiro_e_ultimo_dia_do_mes(ano, mes)
        
        # Adiciona o mês e ano na célula A13
        aba_ativa['A13'] = f"{self.obter_nome_mes(mes)}: {primeiro_dia} à {ultimo_dia}".upper()

        # Ajusta os dados na aba ativa
        self.ajustar_data(mes, ano, aba_ativa, recesso_inicio, recesso_fim)

        rf_numeros = re.sub(r'\D', '', rf)  # Remove pontos e barra do RF, mantendo apenas números

        # Garante que o novo arquivo é salvo com a extensão correta
        arquivo_nome = f"Planilha_{mes:02}_{ano}_{nome_servidor}_{rf_numeros}.xlsm"
        wb_origem.save(os.path.join(diretorio, arquivo_nome))
        print(f"Planilha de horas para {nome_servidor} (RF: {rf_numeros}) de: {mes}/{ano} gerada com sucesso!")
 
    # Função para validar entrada de dados
    def obter_inteiro_valido(self, prompt, minimo, maximo):
        while True:
            try:
                valor = int(input(prompt))
                if minimo <= valor <= maximo:
                    return valor
                else:
                    print(f"Por favor, insira um valor entre {minimo} e {maximo}.")
            except ValueError:
                print("Entrada inválida. Por favor, insira um número inteiro válido.")

    def obter_resposta_sim_nao(self, prompt):
        while True:
            resposta = input(prompt).strip().lower()
            if resposta in ['s', 'n']:
                return resposta
            else:
                print("Entrada inválida. Por favor, digite 's' para sim ou 'n' para não.")
